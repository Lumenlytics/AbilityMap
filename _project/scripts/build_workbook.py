"""
Build the full AbilityMap workbook: a Summary tab + one tab per class.

    python build_workbook.py            # all 13 classes found in ../data/
    python build_workbook.py WARRIOR    # just one

Reads  ../data/<class>_rows.json
Writes ../output/AbilityMap.xlsx

Layout per class tab: each castable Ability is a shaded header row, with the
talents that AUGMENT it nested beneath as "-> Modifier" rows, then a
"General / Passive" section. Colour marks confidence, not decoration.
"""
import json, os, sys, glob
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'data')
OUT = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'output',
                   'AbilityMap.xlsx')

DISPLAY = {
    'deathknight': 'Death Knight', 'demonhunter': 'Demon Hunter', 'druid': 'Druid',
    'evoker': 'Evoker', 'hunter': 'Hunter', 'mage': 'Mage', 'monk': 'Monk',
    'paladin': 'Paladin', 'priest': 'Priest', 'rogue': 'Rogue', 'shaman': 'Shaman',
    'warlock': 'Warlock', 'warrior': 'Warrior',
}

FONT = 'Calibri'
HDR_FILL = PatternFill('solid', fgColor='2F4F6F')
HDR_FONT = Font(name=FONT, bold=True, color='FFFFFF', size=10)
ABILITY_FILL = PatternFill('solid', fgColor='DCE6F1')
ABILITY_FONT = Font(name=FONT, bold=True, size=10)
MOD_FONT = Font(name=FONT, size=10, italic=True, color='555555')
PASS_FONT = Font(name=FONT, size=10, color='666666')
SECTION_FILL = PatternFill('solid', fgColor='F2F2F2')

VERIFIED_FILL = PatternFill('solid', fgColor='C6EFCE')   # green  = hand-checked
SEED_FILL = PatternFill('solid', fgColor='FFF2CC')       # amber  = SimC-seeded
WARN_FILL = PatternFill('solid', fgColor='FCE4D6')       # orange = needs a look

THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (header, json key, width, align)
COLS = [
    ('Ability Group', 'Group', 22, 'left'),
    ('Kind', 'Kind', 13, 'left'),
    ('Name', 'Name', 26, 'left'),
    ('Spell ID', 'SpellID', 10, 'center'),
    ('Also IDs', 'AltIDs', 13, 'center'),
    ('Specs', 'Specs', 20, 'left'),
    ('Active?', 'Type', 9, 'center'),
    ('Cooldown', '_cd', 10, 'center'),
    ('Charges', 'Charges', 8, 'center'),
    ('Aura ID', '_auraid', 10, 'center'),
    ('Aura', '_aurakind', 9, 'center'),
    ('Lands On', 'AuraUnit', 10, 'center'),
    ('Aura Lasts', '_auradur', 11, 'center'),
    ('Renamed To', 'Override', 11, 'center'),
    ('Passive Changes', 'Affects', 16, 'left'),
    ('Confidence', 'Confidence', 17, 'left'),
    ('Caveat', 'Note', 40, 'left'),
    ('Description (from game data)', 'Desc', 72, 'left'),
]


def derive(r):
    """Flatten buff/debuff into one aura column and format the display values."""
    buff, debuff = r.get('BuffID', ''), r.get('DebuffID', '')
    r['_auraid'] = buff or debuff or ''
    r['_aurakind'] = 'buff' if buff else ('debuff' if debuff else '')
    if r.get('AuraPermanent'):
        r['_auradur'] = 'permanent'
    elif r.get('AuraDuration') not in ('', None):
        r['_auradur'] = f"{r['AuraDuration']}s"
    else:
        r['_auradur'] = ''
    cd = r.get('Cooldown_s')
    r['_cd'] = f"{cd}s" if cd not in ('', None) else ''
    return r


def write_class(wb, lower, rows):
    ws = wb.create_sheet(DISPLAY.get(lower, lower.title()))
    for c, (h, _, w, _a) in enumerate(COLS, 1):
        cell = ws.cell(1, c, h)
        cell.fill, cell.font, cell.border = HDR_FILL, HDR_FONT, BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'D2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLS))}1"
    ws.row_dimensions[1].height = 30

    r_i = 2
    for row in rows:
        derive(row)
        kind = row['Kind'].strip()
        is_ability = kind == 'Ability'
        is_mod = 'Modifier' in kind
        conf = str(row.get('Confidence', ''))
        for c, (h, key, w, align) in enumerate(COLS, 1):
            val = row.get(key, '')
            if key == 'Kind':
                val = {'Ability': 'Ability', 'Passive': 'Passive'}.get(kind, '  ↳ Modifier')
            if key == 'Type':
                val = 'Active' if row.get('Type') == 'Active' else '-'
            cell = ws.cell(r_i, c, val)
            cell.border = BORDER
            cell.alignment = Alignment(horizontal=align, vertical='top',
                                       wrap_text=(key in ('Desc', 'Note')))
            cell.font = ABILITY_FONT if is_ability else (MOD_FONT if is_mod else PASS_FONT)
            if is_ability and key != 'Desc':
                cell.fill = ABILITY_FILL
            elif row['Group'] == 'General / Passive' and key == 'Group':
                cell.fill = SECTION_FILL
        # confidence colouring on the aura + confidence cells
        conf_col = [i for i, c in enumerate(COLS, 1) if c[1] == 'Confidence'][0]
        aura_col = [i for i, c in enumerate(COLS, 1) if c[1] == '_auraid'][0]
        if conf == 'Verified':
            ws.cell(r_i, conf_col).fill = VERIFIED_FILL
            ws.cell(r_i, aura_col).fill = VERIFIED_FILL
        elif conf.startswith('SimC'):
            ws.cell(r_i, conf_col).fill = SEED_FILL
            if row['_auraid']:
                ws.cell(r_i, aura_col).fill = SEED_FILL
        if row.get('Note'):
            note_col = [i for i, c in enumerate(COLS, 1) if c[1] == 'Note'][0]
            ws.cell(r_i, note_col).fill = WARN_FILL
        r_i += 1
    return ws


def summary(wb, stats):
    ws = wb.create_sheet('Summary', 0)
    ws.column_dimensions['A'].width = 22
    for col, w in zip('BCDEFG', (11, 11, 13, 13, 15, 15)):
        ws.column_dimensions[col].width = w

    def put(r, c, v, bold=False, size=11, fill=None, align='left'):
        cell = ws.cell(r, c, v)
        cell.font = Font(name=FONT, bold=bold, size=size)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        if fill:
            cell.fill = fill
        return cell

    put(1, 1, 'AbilityMap — WoW Midnight 12.0.7 (build 68453)', True, 16)
    put(2, 1, 'Every ability mapped to the buff/debuff aura it applies, with cooldowns, '
              'durations and the talents that modify it.', False, 10)

    hdr = ['Class', 'Rows', 'Abilities', 'With Aura', 'Player Auras', 'Target Auras',
           'Timing Passives']
    for c, h in enumerate(hdr, 1):
        cell = put(4, c, h, True, 10, HDR_FILL, 'center')
        cell.font = HDR_FONT
        cell.border = BORDER

    r = 5
    tot = [0] * 6
    for name, s in stats:
        put(r, 1, name, True, 10)
        vals = [s['rows'], s['abilities'], s['auras'], s['player'], s['target'], s['timing']]
        for i, v in enumerate(vals):
            put(r, i + 2, v, False, 10, None, 'center')
            tot[i] += v
        r += 1
    put(r, 1, 'TOTAL', True, 10, SECTION_FILL)
    for i, v in enumerate(tot):
        put(r, i + 2, v, True, 10, SECTION_FILL, 'center')

    r += 2
    put(r, 1, 'How to read a class tab', True, 13); r += 1
    for line in [
        'Rows are grouped by ability. A shaded bold row is a castable Ability; the italic',
        '"↳ Modifier" rows beneath it are the talents that augment that ability. Passives that',
        'do not modify a specific ability sit at the bottom under "General / Passive".',
        '',
        'Aura ID — the buff/debuff the ability applies. This is usually a DIFFERENT spell ID',
        '   than the cast: Shield Block 2565 applies buff 132404; Mortal Strike 12294 applies',
        '   debuff Mortal Wounds 213667. Never assume they match.',
        'Lands On — player or target. Decides which unit an addon watches for that aura.',
        'Renamed To — a talent replaces this ability with that spell ID (Judgment 20271 -> 275779).',
        'Passive Changes — what a passive alters. duration / cooldown / charges affect an',
        '   ability\'s timing; "damage" ones do not.',
        'Also IDs — other spell IDs for the same ability (spec or rank variants), merged into one row.',
    ]:
        put(r, 1, line, False, 10); r += 1

    r += 1
    put(r, 1, 'Confidence — colour coded', True, 13); r += 1
    put(r, 1, 'Verified', True, 10, VERIFIED_FILL)
    put(r, 2, 'Hand-checked against Wowhead. Currently Warrior only (38 auras).', False, 10); r += 1
    put(r, 1, 'SimC (...)', True, 10, SEED_FILL)
    put(r, 2, 'From SimulationCraft\'s extraction of Blizzard client data. Scored 38/38 correct '
              '(28/28 durations) against the verified Warrior set — strong, but not independently checked per class.',
       False, 10); r += 1
    put(r, 1, 'Captured', True, 10)
    put(r, 2, 'Ability read from the live client, but no aura identified.', False, 10); r += 1
    put(r, 1, 'Caveat column', True, 10, WARN_FILL)
    put(r, 2, 'A known problem with that row — read it before relying on the number.', False, 10); r += 2

    put(r, 1, 'Known limitations', True, 13); r += 1
    for line in [
        'A cooldown of 1s or less on an ability with no charge count is usually the inter-charge',
        '   lockout, not the real recharge time. Those rows carry a caveat.',
        'Charge-ability cooldowns read from the live client include haste and talent reductions,',
        '   so they sit below the true base value (Shield Block reads 13.6s, base is 16s).',
        'PvP talents are missing for classes not played — no cross-class API exists for them.',
        'Aura durations are base values: no haste scaling, talent extensions or pandemic behaviour.',
    ]:
        put(r, 1, line, False, 10); r += 1

    r += 1
    put(r, 1, 'Sources: live client capture via the CCXMap addon; SimulationCraft SpellDataDump '
              '(GPL-3.0) for aura links and baseline spells.', False, 9)


def main():
    only = sys.argv[1].lower() if len(sys.argv) > 1 else None
    wb = Workbook()
    wb.remove(wb.active)
    stats = []
    files = sorted(glob.glob(os.path.join(DATA, '*_rows.json')))
    for path in files:
        lower = os.path.basename(path).replace('_rows.json', '')
        if only and lower != only:
            continue
        rows = json.load(open(path, encoding='utf-8'))
        write_class(wb, lower, rows)
        stats.append((DISPLAY.get(lower, lower.title()), {
            'rows': len(rows),
            'abilities': sum(1 for r in rows if r['Kind'].strip() == 'Ability'),
            'auras': sum(1 for r in rows if r.get('BuffID') or r.get('DebuffID')),
            'player': sum(1 for r in rows if r.get('AuraUnit') == 'player'
                          and r.get('Type') == 'Active'),
            'target': sum(1 for r in rows if r.get('AuraUnit') == 'target'
                          and r.get('Type') == 'Active'),
            'timing': sum(1 for r in rows if r.get('Affects') and
                          any(t in r['Affects'] for t in ('duration', 'cooldown', 'charges'))),
        }))
    summary(wb, stats)
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"wrote {OUT}")
    print(f"  {len(stats)} class tabs + Summary, "
          f"{sum(s['rows'] for _, s in stats)} rows total")


if __name__ == '__main__':
    main()
